import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
import tempfile
from openpyxl import load_workbook
from difflib import SequenceMatcher
from io import BytesIO
import gspread
from google.oauth2.service_account import Credentials
import requests
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbziZ27mG690ZT02YN1LqbvWJLZ-rprnHK9qmXDDXcTvQVmnB-Phpm0J4DKjsg6Ts07xJQ/exec"

def send_to_google_sheet(sheet_name, row):
    payload = {
        "sheet_name": sheet_name,
        "row": row
    }

    response = requests.post(GOOGLE_SCRIPT_URL, json=payload)

    if response.status_code == 200:
        return response.json()
    else:
        return {"status": "error", "message": response.text}

HEADER_PATH = "header.png"
TEMPLATE1_PATH = "templates.xlsx"
TEMPLATE2_PATH = "templates2.xlsx"

DATA_PATH = "Jidhafs.xlsx"
st.set_page_config(page_title="نظام الزيارات الصفية", layout="wide")

st.markdown("""
<style>
.stApp {
    background: linear-gradient(135deg, #f7f9ff, #fff8f1);
    direction: rtl;
}
.big-title {
    text-align: center;
    font-size: 50px;
    font-weight: 900;
    color: #1f2937;
    margin-bottom: 30px;
}
.card {
    background: white;
    border-radius: 22px;
    padding: 22px;
    box-shadow: 0 6px 18px rgba(0,0,0,0.07);
    border-right: 7px solid #6366f1;
    text-align: right;
}
.card-title {
    font-size: 17px;
    color: #6b7280;
    font-weight: 700;
}
.card-value {
    font-size: 42px;
    color: #111827;
    font-weight: 900;
}
.section-title {
    text-align: right;
    font-size: 28px;
    font-weight: 900;
    margin-top: 30px;
    margin-bottom: 15px;
    color: #1f2937;
}
</style>
""", unsafe_allow_html=True)

st.image(HEADER_PATH, use_container_width=True)
st.markdown('<div class="big-title">📊 نظام الزيارات الصفية</div>', unsafe_allow_html=True)

def load_data():
    df = pd.read_excel(DATA_PATH, sheet_name="Main")
    df.columns = [str(c).strip() for c in df.columns]
    return df

df = load_data()
st.success(f"✅ تم تحميل البيانات: {len(df)} سجل")

def clean_text(x):
    return (
        str(x)
        .replace("\n", "")
        .replace("\r", "")
        .replace(" ", "")
        .replace("ـ", "")
        .replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ى", "ي")
        .strip()
    )

def find_col(keywords):
    for col in df.columns:
        col_clean = clean_text(col)
        for key in keywords:
            if clean_text(key) in col_clean:
                return col

    st.error(f"❌ لم أجد عمود يحتوي على: {keywords}")
    st.write("الأعمدة الموجودة:")
    st.write(df.columns.tolist())
    st.stop()

year_col = find_col(["السنة الدراسية", "السنة"])
term_col = find_col(["الفصل الدراسي", "الفصل"])
dept_col = find_col(["الأقسام الأكاديمية", "القسم الأكاديمي", "القسم"])
teacher_col = find_col(["اسم المعلمة", "المعلمة"])
visitor_col = find_col(["الزائر"])
month_col = find_col(["الشهر"])

judgment_col = find_col([
    "الحكم العام",
    "الحكم العام ",
    "الحكم",
    "التقييم العام"
])

df[judgment_col] = df[judgment_col].fillna("تقييم ذاتي")
df[judgment_col] = df[judgment_col].astype(str).str.strip()
df[judgment_col] = df[judgment_col].replace(["nan", ""], "تقييم ذاتي")

self_eval_col = find_col(["حكم للتقييم الذاتي"])

for col in [year_col, term_col, dept_col, teacher_col, visitor_col, month_col, judgment_col, self_eval_col]:
    df[col] = df[col].astype(str).str.strip()
# =======================
# صلاحيات دخول الأقسام
# =======================
dept_passwords = {
    "admin1825": "الكل",
    "Arab1111": "قسم اللغة العربية",
    "Math2222": "قسم الرياضيات",
    "Sc3333": "قسم العلوم",
    "Islamic4444": "قسم التربية الإسلامية",
    "Ict5555": "قسم الحاسب الآلي",
    "Eng6666": "قسم اللغة الإنجليزية",
    "Social7777": "قسم المواد الاجتماعية",
    "Art8888": "قسم التربية الفنية",
    "Sport9999": "قسم التربية البدنية",
    "Com1010": "قسم المواد التجارية",
    "Family2020": "قسم التربية الأسرية"
}

st.sidebar.markdown("## 🔐 الدخول")
# حالة تسجيل الدخول
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
    st.session_state["allowed_dept"] = None

# إذا مو مسجل دخول
if not st.session_state["logged_in"]:
    site_password = st.sidebar.text_input("أدخلي الرقم السري", type="password")

    if site_password:
        if site_password in dept_passwords:
            st.session_state["logged_in"] = True
            st.session_state["allowed_dept"] = dept_passwords[site_password]
            st.rerun()
        else:
            st.sidebar.error("❌ الرقم السري غير صحيح")

    st.stop()

# إذا مسجل دخول
allowed_dept = st.session_state["allowed_dept"]
st.sidebar.markdown("---")

if st.sidebar.button("🚪 تسجيل الخروج"):
    st.session_state["logged_in"] = False
    st.session_state["allowed_dept"] = None
    st.rerun()
# =======================
# الفلاتر
# =======================
st.sidebar.markdown("## 🎯 الفلاتر")

def clean_value(x):
    return str(x).strip()

for col in [year_col, term_col, dept_col, teacher_col, visitor_col, month_col, judgment_col]:
    df[col] = df[col].apply(clean_value)

def dropdown(label, dataframe, col, key):
    values = sorted([
        v for v in dataframe[col].dropna().astype(str).unique()
        if v and v.lower() != "nan" and v != ""
    ])
    return st.sidebar.selectbox(label, ["الكل"] + values, key=key)

filtered = df.copy()

selected_year = dropdown("السنة الدراسية", filtered, year_col, "filter_year")
if selected_year != "الكل":
    filtered = filtered[filtered[year_col] == selected_year]

selected_term = dropdown("الفصل الدراسي", filtered, term_col, "filter_term")
if selected_term != "الكل":
    filtered = filtered[filtered[term_col] == selected_term]

def norm(x):
    return str(x).strip().replace("ـ", "")

if allowed_dept == "الكل":
    selected_dept = dropdown("القسم الأكاديمي", filtered, dept_col, "filter_dept")
    if selected_dept != "الكل":
        filtered = filtered[filtered[dept_col].apply(norm) == norm(selected_dept)]
else:
    selected_dept = allowed_dept
    filtered = filtered[filtered[dept_col].apply(norm) == norm(allowed_dept)]
    st.sidebar.success(f"القسم: {allowed_dept}")

selected_teacher = dropdown("اسم المعلمة", filtered, teacher_col, "filter_teacher")
if selected_teacher != "الكل":
    filtered = filtered[filtered[teacher_col] == selected_teacher]

selected_visitor = dropdown("الزائر", filtered, visitor_col, "filter_visitor")
if selected_visitor != "الكل":
    filtered = filtered[filtered[visitor_col] == selected_visitor]

selected_month = dropdown("الشهر", filtered, month_col, "filter_month")
if selected_month != "الكل":
    filtered = filtered[filtered[month_col] == selected_month]

judgment_counts = filtered[judgment_col].value_counts()

# =======================
# المؤشرات العامة
# =======================
st.markdown('<div class="section-title">📌 المؤشرات العامة</div>', unsafe_allow_html=True)

c1, c2, c3, c4 = st.columns(4)

metrics = [
    ("عدد الزيارات", len(filtered), "#6366f1"),
    ("عدد المعلمات", filtered[teacher_col].dropna().str.strip().nunique(), "#ec4899"),
    ("عدد الأقسام", filtered[dept_col].dropna().str.strip().nunique(), "#10b981"),
    ("عدد الزوار", filtered[visitor_col].dropna().str.strip().nunique(), "#f97316"),
]

for col, (title, value, color) in zip([c1, c2, c3, c4], metrics):
    with col:
        st.markdown(f"""
        <div class="card" style="border-right-color:{color};">
            <div class="card-title">{title}</div>
            <div class="card-value">{value}</div>
        </div>
        """, unsafe_allow_html=True)
        
# =======================
# مؤشرات إضافية
# =======================

# أكثر زائر قام بزيارات
if not filtered.empty:
    top_visitor = filtered[visitor_col].value_counts().idxmax()
    top_visitor_count = filtered[visitor_col].value_counts().max()
else:
    top_visitor = "لا يوجد"
    top_visitor_count = 0

# أكثر قسم تمت زيارته
if not filtered.empty:
    top_dept = filtered[dept_col].value_counts().idxmax()
    top_dept_count = filtered[dept_col].value_counts().max()
else:
    top_dept = "لا يوجد"
    top_dept_count = 0

st.markdown('<div class="section-title">🔎 مؤشرات الزيارات</div>', unsafe_allow_html=True)

v1, v2 = st.columns(2)

with v1:
    st.markdown(f"""
    <div class="card" style="border-right-color:#8b5cf6;">
        <div class="card-title">أكثر زائر قام بزيارات</div>
        <div class="card-value" style="font-size:28px;">{top_visitor}</div>
        <div style="font-size:20px;font-weight:700;color:#6b7280;">عدد الزيارات: {top_visitor_count}</div>
    </div>
    """, unsafe_allow_html=True)

with v2:
    st.markdown(f"""
    <div class="card" style="border-right-color:#06b6d4;">
        <div class="card-title">أكثر قسم تمت زيارته</div>
        <div class="card-value" style="font-size:28px;">{top_dept}</div>
        <div style="font-size:20px;font-weight:700;color:#6b7280;">عدد الزيارات: {top_dept_count}</div>
    </div>
    """, unsafe_allow_html=True)

# =======================
# الأحكام
# =======================
st.markdown('<div class="section-title">📊 الأحكام</div>', unsafe_allow_html=True)

judgment_colors = {
    "يتجاوز التوقعات بكثير": "#B7E1A1",
    "يتجاوز التوقعات": "#BFE9F7",
    "يفي بالتوقعات": "#FFF59D",
    "يفي بالتوقعات جزئياً": "#EAC2E8",
    "يفي بالتوقعات جزئيا": "#EAC2E8",
    "تقييم ذاتي": "#D7CCC8"
}

total_judgments = (
    judgment_counts.get("يتجاوز التوقعات بكثير", 0) +
    judgment_counts.get("يتجاوز التوقعات", 0) +
    judgment_counts.get("يفي بالتوقعات", 0) +
    judgment_counts.get("يفي بالتوقعات جزئياً", 0) +
    judgment_counts.get("يفي بالتوقعات جزئيا", 0)
)

score = (
    judgment_counts.get("يتجاوز التوقعات بكثير", 0) * 4 +
    judgment_counts.get("يتجاوز التوقعات", 0) * 3 +
    judgment_counts.get("يفي بالتوقعات", 0) * 2 +
    (judgment_counts.get("يفي بالتوقعات جزئياً", 0) + judgment_counts.get("يفي بالتوقعات جزئيا", 0)) * 1
)

performance_index = score / (total_judgments * 4) if total_judgments > 0 else 0
total_visits = len(filtered)

j1, j2, j3, j4, j5 = st.columns(5)

judgment_cards = [
    ("مؤشر الأداء", f"{performance_index*100:.1f}%", "#FCE4D6"),
    ("يتجاوز التوقعات بكثير", judgment_counts.get("يتجاوز التوقعات بكثير", 0), "#B7E1A1"),
    ("يتجاوز التوقعات", judgment_counts.get("يتجاوز التوقعات", 0), "#BFE9F7"),
    ("يفي بالتوقعات", judgment_counts.get("يفي بالتوقعات", 0), "#FFF59D"),
    ("يفي بالتوقعات جزئياً", judgment_counts.get("يفي بالتوقعات جزئياً", 0) + judgment_counts.get("يفي بالتوقعات جزئيا", 0), "#EAC2E8"),
]

for col, (title, value, bg_color) in zip([j1, j2, j3, j4, j5], judgment_cards):
    with col:
        st.markdown(f"""
        <div style="
            background:{bg_color};
            border-radius:18px;
            padding:18px;
            text-align:center;
            box-shadow:0 5px 14px rgba(0,0,0,0.08);
            border:1px solid #ddd;
            min-height:130px;
        ">
            <div style="font-size:16px;font-weight:800;color:#111827;">{title}</div>
            <div style="font-size:36px;font-weight:900;color:#000;margin-top:8px;">{value}</div>
        </div>
        """, unsafe_allow_html=True)

if total_visits > 0 and not judgment_counts.empty:
    overall_judgment = judgment_counts.idxmax()
else:
    overall_judgment = "لا توجد بيانات"

overall_color = judgment_colors.get(overall_judgment, "#F3F4F6")

st.markdown(f"""
<div style="margin-top:20px;display:flex;justify-content:center;">
    <div style="
        display:flex;
        width:520px;
        border:1px solid #333;
        font-size:20px;
        font-weight:900;
        text-align:center;
        border-radius:8px;
        overflow:hidden;
    ">
        <div style="width:50%;background:#f3f4f6;padding:14px;">الحكم العام</div>
        <div style="width:50%;background:{overall_color};padding:14px;">{overall_judgment}</div>
    </div>
</div>
""", unsafe_allow_html=True)

judgment_chart = pd.DataFrame({
    "الحكم": [
        "يتجاوز التوقعات بكثير",
        "يتجاوز التوقعات",
        "يفي بالتوقعات",
        "يفي بالتوقعات جزئياً"
    ],
    "العدد": [
        judgment_counts.get("يتجاوز التوقعات بكثير", 0),
        judgment_counts.get("يتجاوز التوقعات", 0),
        judgment_counts.get("يفي بالتوقعات", 0),
        judgment_counts.get("يفي بالتوقعات جزئياً", 0) + judgment_counts.get("يفي بالتوقعات جزئيا", 0)
    ]
})

fig_judgment = px.bar(
    judgment_chart,
    x="الحكم",
    y="العدد",
    text="العدد",
    title="توزيع الأحكام",
    color="الحكم",
    color_discrete_map=judgment_colors
)

fig_judgment.update_traces(textposition="outside")
fig_judgment.update_layout(height=470, title_x=0.5, font=dict(size=16), showlegend=False)
st.plotly_chart(fig_judgment, use_container_width=True, key="judgment_chart")

# =======================
# تحليل البنود والأقسام
# =======================
st.markdown('<div class="section-title">📋 تحليل البنود وترتيب الأقسام</div>', unsafe_allow_html=True)

score_map = {
    "يتجاوز التوقعات بكثير": 4,
    "يتجاوز التوقعات": 3,
    "يفي بالتوقعات": 2,
    "يفي بالتوقعات جزئياً": 1,
    "تقييم ذاتي": 2
}

analysis_df = filtered.copy()
analysis_df["درجة الحكم"] = analysis_df[judgment_col].map(score_map).fillna(0)

# تحديد البنود من O إلى AF حسب ترتيب Excel
# ترتيب البنود حسب التمبليت:
# أولاً AA و AB، ثم من O إلى Z، ثم من AC إلى AF
item_cols = (
    list(filtered.columns[26:28]) +   # AA, AB
    list(filtered.columns[14:26]) +   # O إلى Z
    list(filtered.columns[28:32])     # AC إلى AF
)

dept_analysis = (
    analysis_df
    .groupby(dept_col)
    .agg(
        عدد_الزيارات=("درجة الحكم", "count"),
        عدد_المعلمات=(teacher_col, "nunique"),
        متوسط_الأداء=("درجة الحكم", "mean")
    )
    .reset_index()
)

dept_analysis["مؤشر الأداء"] = (dept_analysis["متوسط_الأداء"] / 4 * 100).round(1)
dept_analysis = dept_analysis.sort_values("مؤشر الأداء", ascending=False)

def get_dept_judgment(score):
    if score >= 90:
        return "يتجاوز التوقعات بكثير"
    elif score >= 80:
        return "يتجاوز التوقعات"
    elif score >= 70:
        return "يفي بالتوقعات"
    else:
        return "يفي بالتوقعات جزئياً"

dept_analysis["الحكم"] = dept_analysis["مؤشر الأداء"].apply(get_dept_judgment)

dept_color_map = {
    "يتجاوز التوقعات بكثير": "#B7E1A1",
    "يتجاوز التوقعات": "#BFE9F7",
    "يفي بالتوقعات": "#FFF59D",
    "يفي بالتوقعات جزئياً": "#EAC2E8"
}

fig_dept_rank = px.bar(
    dept_analysis,
    x=dept_col,
    y="مؤشر الأداء",
    text="مؤشر الأداء",
    color="الحكم",
    color_discrete_map=dept_color_map,
    title="ترتيب الأقسام حسب مؤشر الأداء",
    custom_data=["الحكم", "عدد_الزيارات", "عدد_المعلمات"]
)

fig_dept_rank.update_traces(
    hovertemplate=
    "<b>القسم:</b> %{x}<br>" +
    "<b>مؤشر الأداء:</b> %{y}%<br>" +
    "<b>الحكم:</b> %{customdata[0]}<br>" +
    "<b>عدد الزيارات:</b> %{customdata[1]}<br>" +
    "<b>عدد المعلمات:</b> %{customdata[2]}" +
    "<extra></extra>"
)

fig_dept_rank.update_layout(height=450, title_x=0.5)
st.plotly_chart(fig_dept_rank, use_container_width=True, key="dept_rank_chart")

# =======================
# تحليل البنود حسب الأحكام
# =======================
st.markdown('<div class="section-title">📊 تحليل البنود حسب الأحكام</div>', unsafe_allow_html=True)

items_rows = []

for item in item_cols:
    values = filtered[item].dropna().astype(str).str.strip()
    values = values.replace("يفي بالتوقعات جزئيا", "يفي بالتوقعات جزئياً")

    for judgment in [
        "يتجاوز التوقعات بكثير",
        "يتجاوز التوقعات",
        "يفي بالتوقعات",
        "يفي بالتوقعات جزئياً"
    ]:
        items_rows.append({
            "البند": item,
            "الحكم": judgment,
            "العدد": int((values == judgment).sum())
        })

items_plot = pd.DataFrame(items_rows)


fig_items = px.bar(
    items_plot,
    y="البند",
    x="العدد",
    color="الحكم",
    color_discrete_map=judgment_colors,
    barmode="stack",
    orientation="h",
    title="📊 توزيع الأحكام حسب البنود"
)

fig_items.update_layout(
    height=700,
    title_x=0.5,
    yaxis=dict(
        autorange="reversed"   # 👈 عكس البنود
    )
)

st.plotly_chart(fig_items, use_container_width=True)

# =======================
# تحليل الأحكام حسب المعلمة
# =======================
st.markdown('<div class="section-title">👩‍🏫 توزيع الأحكام حسب المعلمة</div>', unsafe_allow_html=True)

teacher_plot = (
    filtered
    .groupby([teacher_col, judgment_col])
    .size()
    .reset_index(name="العدد")
)

# ترتيب أسماء المعلمات حسب إجمالي الزيارات
teacher_order = (
    filtered[teacher_col]
    .value_counts()
    .index
    .tolist()
)

fig_teacher_judgments = px.bar(
    teacher_plot,
    y=teacher_col,
    x="العدد",
    color=judgment_col,
    color_discrete_map=judgment_colors,
    barmode="stack",
    orientation="h",
    title="📊 توزيع الأحكام حسب المعلمة",
    category_orders={teacher_col: teacher_order}
)

fig_teacher_judgments.update_layout(
    height=max(600, len(teacher_order) * 28),
    title_x=0.5,
    yaxis=dict(
        autorange="reversed",
        side="right",
        tickfont=dict(size=11)
    ),
    xaxis=dict(
        title="عدد الأحكام",
        autorange="reversed"
    ),
    legend=dict(
        orientation="h",
        yanchor="bottom",
        y=1.04,
        xanchor="center",
        x=0.5
    ),
    margin=dict(r=280, l=40, t=120, b=40)
)

st.plotly_chart(fig_teacher_judgments, use_container_width=True, key="teacher_judgments_chart")

# =======================
# ملخص الفلترة
# =======================
st.markdown('<div class="section-title">🧾 ملخص الفلترة الحالية</div>', unsafe_allow_html=True)

st.info(
    f"""
    السنة الدراسية: {selected_year}  
    الفصل الدراسي: {selected_term}  
    القسم الأكاديمي: {selected_dept}  
    اسم المعلمة: {selected_teacher}  
    الزائر: {selected_visitor}  
    الشهر: {selected_month}
    """
)

# =======================
# الطباعة
# =======================
st.markdown('<div class="section-title">🖨️ طباعة التقرير</div>', unsafe_allow_html=True)

template_choice = st.selectbox(
    "اختاري ملف التمبليت",
    ["templates.xlsx", "templates2.xlsx"]
)

template_path = TEMPLATE1_PATH if template_choice == "templates.xlsx" else TEMPLATE2_PATH

def write_cell(ws, cell, value):
    for merged in ws.merged_cells.ranges:
        if cell in merged:
            ws[merged.start_cell.coordinate] = value
            return
    ws[cell] = value

def clean_match_text(x):
    return str(x).replace(" ", "").replace("\n", "").replace("ـ", "").strip()

def find_matching_column(criteria_text, columns):
    target = clean_match_text(criteria_text)
    best_col = None
    best_score = 0

    for col in columns:
        score = SequenceMatcher(None, target, clean_match_text(col)).ratio()
        if score > best_score:
            best_score = score
            best_col = col

    return best_col if best_score >= 0.35 else None

def get_first_text(df, keywords):
    for col in df.columns:
        if any(k in str(col) for k in keywords):
            vals = df[col].dropna().astype(str).str.strip()
            vals = vals[vals != ""]
            if len(vals) > 0:
                return "\n".join(vals.unique()[:5])
    return ""

def fill_template_streamlit():
    selected_sheet = st.session_state["selected_print_sheet"]
    report_df = filtered.copy()

    wb = load_workbook(template_path)
    ws = wb[selected_sheet]

    # حذف كل الشيتات ما عدا الشيت المختار
    for sheet in wb.sheetnames[:]:
        if sheet != selected_sheet:
            del wb[sheet]

    # تعبئة بيانات التقرير
    ws["C3"] = selected_year
    ws["F3"] = selected_term
    ws["C4"] = selected_dept
    ws["F4"] = selected_visitor
    ws["C5"] = selected_teacher
    ws["F5"] = len(report_df)
    ws["F6"] = selected_month

    total_counts = {
        "يتجاوز التوقعات بكثير": 0,
        "يتجاوز التوقعات": 0,
        "يفي بالتوقعات": 0,
        "يفي بالتوقعات جزئياً": 0,
    }

    def normalize_judgment(x):
        x = str(x).strip()
        if x == "يفي بالتوقعات جزئيا":
            return "يفي بالتوقعات جزئياً"
        return x

    judgment_to_col = {
        "يفي بالتوقعات جزئياً": "D",
        "يفي بالتوقعات": "E",
        "يتجاوز التوقعات": "F",
        "يتجاوز التوقعات بكثير": "G",
    }

    # البحث عن صف العناوين وصف المجموع
    header_row = None
    total_row = None

    for row in range(1, ws.max_row + 1):
        c_val = str(ws[f"C{row}"].value).strip()
        b_val = str(ws[f"B{row}"].value).strip()

        if "المعايير" in c_val or "الرقم" in b_val:
            header_row = row

        if "المجموع" in c_val:
            total_row = row
            break

    if header_row:
        ws[f"D{header_row}"] = "يفي بالتوقعات جزئياً"
        ws[f"E{header_row}"] = "يفي بالتوقعات"
        ws[f"F{header_row}"] = "يتجاوز التوقعات"
        ws[f"G{header_row}"] = "يتجاوز التوقعات بكثير"

    start_row = header_row + 1 if header_row else 8
    end_row = total_row if total_row else ws.max_row + 1

    is_individual_report = "الفردية" in selected_sheet or "فردية" in selected_sheet

    # ترتيب البنود حسب التمبليت:
    # AA و AB أولاً، ثم O إلى Z، ثم AC إلى AF
    item_cols = (
        list(report_df.columns[26:28]) +   # AA, AB
        list(report_df.columns[14:26]) +   # O إلى Z
        list(report_df.columns[28:32])     # AC إلى AF
    )

    for i, item_col in enumerate(item_cols):
        row = start_row + i

        if row >= end_row:
            break

        # تنظيف خلايا الأحكام
        for col_letter in ["D", "E", "F", "G"]:
            ws[f"{col_letter}{row}"] = ""

        values = report_df[item_col].dropna().astype(str).str.strip()
        values = values.apply(normalize_judgment)

        if is_individual_report:
            if len(values) == 0:
                continue

            judgment_value = values.iloc[0]
            target_col = judgment_to_col.get(judgment_value)

            if target_col:
                ws[f"{target_col}{row}"] = "✓"

            if judgment_value in total_counts:
                total_counts[judgment_value] += 1

        else:
            counts = {
                "يتجاوز التوقعات بكثير": int((values == "يتجاوز التوقعات بكثير").sum()),
                "يتجاوز التوقعات": int((values == "يتجاوز التوقعات").sum()),
                "يفي بالتوقعات": int((values == "يفي بالتوقعات").sum()),
                "يفي بالتوقعات جزئياً": int((values == "يفي بالتوقعات جزئياً").sum()),
            }

            ws[f"D{row}"] = counts["يفي بالتوقعات جزئياً"]
            ws[f"E{row}"] = counts["يفي بالتوقعات"]
            ws[f"F{row}"] = counts["يتجاوز التوقعات"]
            ws[f"G{row}"] = counts["يتجاوز التوقعات بكثير"]

            for key in total_counts:
                total_counts[key] += counts[key]

    # المجموع
    if total_row:
        ws[f"D{total_row}"] = total_counts["يفي بالتوقعات جزئياً"]
        ws[f"E{total_row}"] = total_counts["يفي بالتوقعات"]
        ws[f"F{total_row}"] = total_counts["يتجاوز التوقعات"]
        ws[f"G{total_row}"] = total_counts["يتجاوز التوقعات بكثير"]

    # الحكم العام
    total = sum(total_counts.values())
    score = (
        total_counts["يتجاوز التوقعات بكثير"] * 4 +
        total_counts["يتجاوز التوقعات"] * 3 +
        total_counts["يفي بالتوقعات"] * 2 +
        total_counts["يفي بالتوقعات جزئياً"] * 1
    )

    performance = score / (total * 4) if total > 0 else 0

    if performance >= 0.85:
        general_judgment = "يتجاوز التوقعات بكثير"
    elif performance >= 0.70:
        general_judgment = "يتجاوز التوقعات"
    elif performance >= 0.50:
        general_judgment = "يفي بالتوقعات"
    else:
        general_judgment = "يفي بالتوقعات جزئياً"

    ws["F7"] = performance
    ws["C7"] = general_judgment

    # النجاحات والجوانب
    successes = get_first_text(report_df, ["نجاحات", "نجاح", "نقاط قوة", "إيجابيات"])
    improvements = get_first_text(report_df, ["جوانب بحاجة", "تطوير", "بحاجة", "توصيات"])

    ws["C29"] = successes
    ws["C31"] = improvements

    # حفظ في الذاكرة بدل ملف محلي
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

wb_preview = load_workbook(template_path, read_only=True)
# اختيار نوع التقرير
sheet_choice = st.selectbox("اختاري نوع التقرير", wb_preview.sheetnames)

# حفظ الاختيار
st.session_state["selected_print_sheet"] = sheet_choice

# زر إنشاء التقرير
if st.button("📄 إنشاء التقرير"):
    excel_output = fill_template_streamlit()

    st.download_button(
        "⬇️ تحميل Excel",
        data=excel_output,
        file_name=f"{sheet_choice}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
        

# =======================
# الفوتر
# =======================
st.markdown("""
<hr style="margin-top:40px;">

<div style="
display:flex;
justify-content:space-between;
font-size:16px;
font-weight:600;
color:#374151;
padding:10px 20px;
">

<div>مديرة المدرسة: أ. خلود يعقوب</div>
<div>المديرة المساعدة: أ. سامية سلمان</div>
<div>تصميم وبرمجة: أ. عفاف حسين</div>

</div>
""", unsafe_allow_html=True)